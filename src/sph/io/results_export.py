from __future__ import annotations

"""
Structured per-step results logging + export (CSV + XLSX).

This module exists to capture runtime metrics *without parsing stdout*.
It is pure observability / I/O: it must not modify simulation state or physics.
"""

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import csv
import json
import math
import subprocess

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class StepMetrics:
    step: int
    t: float
    dt: float
    vmax: float
    rho_min: float
    rho_avg: float
    rho_max: float
    err_avg_pct: float
    p_min: float
    p_avg: float
    p_max: float
    neigh_min: int
    neigh_avg: float
    neigh_max: int


@dataclass(frozen=True)
class VxProfileMetrics:
    step: int
    bins: int
    x_window: float
    vx_mean: list[float]


def _safe_float(v: float) -> float:
    try:
        return float(v)
    except Exception:
        return float("nan")


def _git_commit_hash() -> str | None:
    """
    Best-effort git commit hash for reproducibility.
    Returns None if not available (e.g., not a git checkout).
    """
    try:
        out = subprocess.check_output(["git", "rev-parse", "HEAD"], stderr=subprocess.DEVNULL, text=True)
        return out.strip()
    except Exception:
        return None


class ResultsLogger:
    """
    In-memory collector for per-step metrics with export to CSV/XLSX.
    """

    def __init__(self, *, meta: dict[str, Any] | None = None):
        self.steps: list[StepMetrics] = []
        self.vx_profiles: list[VxProfileMetrics] = []
        self.meta: dict[str, Any] = dict(meta or {})

        # Standard meta keys (best-effort).
        self.meta.setdefault("timestamp_utc", datetime.now(timezone.utc).isoformat())
        self.meta.setdefault("git_commit", _git_commit_hash())

    def log_step(self, m: StepMetrics) -> None:
        self.steps.append(m)

    def log_vxprof(self, m: VxProfileMetrics) -> None:
        self.vx_profiles.append(m)

    def export(
        self,
        out_dir: str | Path,
        *,
        base_name: str = "results",
        formats: Iterable[str] = ("csv", "xlsx"),
    ) -> None:
        out_dir = Path(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        fmt_set = {str(f).lower() for f in formats}
        if "csv" in fmt_set:
            self._export_csv(out_dir, base_name=base_name)
        if "xlsx" in fmt_set:
            self._export_xlsx(out_dir, base_name=base_name)

    def _export_csv(self, out_dir: Path, *, base_name: str) -> None:
        steps_path = out_dir / f"{base_name}_steps.csv"
        vx_path = out_dir / f"{base_name}_vxprof.csv"

        # Steps CSV
        with steps_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(
                [
                    "step",
                    "t",
                    "dt",
                    "vmax",
                    "rho_min",
                    "rho_avg",
                    "rho_max",
                    "err_avg_pct",
                    "p_min",
                    "p_avg",
                    "p_max",
                    "neigh_min",
                    "neigh_avg",
                    "neigh_max",
                ]
            )
            for m in self.steps:
                w.writerow(
                    [
                        int(m.step),
                        _safe_float(m.t),
                        _safe_float(m.dt),
                        _safe_float(m.vmax),
                        _safe_float(m.rho_min),
                        _safe_float(m.rho_avg),
                        _safe_float(m.rho_max),
                        _safe_float(m.err_avg_pct),
                        _safe_float(m.p_min),
                        _safe_float(m.p_avg),
                        _safe_float(m.p_max),
                        int(m.neigh_min),
                        _safe_float(m.neigh_avg),
                        int(m.neigh_max),
                    ]
                )

        # VXPROF CSV
        max_bins = max((int(v.bins) for v in self.vx_profiles), default=0)
        vx_cols = [f"vx_mean_{i}" for i in range(max_bins)]
        with vx_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["step", "bins", "x_window", *vx_cols])
            for v in self.vx_profiles:
                row = [int(v.step), int(v.bins), _safe_float(v.x_window)]
                vx = list(v.vx_mean)
                vx = vx[:max_bins] + [float("nan")] * max(0, max_bins - len(vx))
                row.extend(_safe_float(x) for x in vx)
                w.writerow(row)

    def _export_xlsx(self, out_dir: Path, *, base_name: str) -> None:
        path = out_dir / f"{base_name}.xlsx"
        wb = Workbook()

        # ---- steps sheet
        ws = wb.active
        ws.title = "steps"
        headers = [
            "step",
            "t",
            "dt",
            "vmax",
            "rho_min",
            "rho_avg",
            "rho_max",
            "err_avg_pct",
            "p_min",
            "p_avg",
            "p_max",
            "neigh_min",
            "neigh_avg",
            "neigh_max",
        ]
        ws.append(headers)
        for m in self.steps:
            ws.append(
                [
                    int(m.step),
                    _safe_float(m.t),
                    _safe_float(m.dt),
                    _safe_float(m.vmax),
                    _safe_float(m.rho_min),
                    _safe_float(m.rho_avg),
                    _safe_float(m.rho_max),
                    _safe_float(m.err_avg_pct),
                    _safe_float(m.p_min),
                    _safe_float(m.p_avg),
                    _safe_float(m.p_max),
                    int(m.neigh_min),
                    _safe_float(m.neigh_avg),
                    int(m.neigh_max),
                ]
            )
        ws.freeze_panes = "A2"
        _autosize_worksheet(ws)

        # ---- vx profile sheet
        ws_vx = wb.create_sheet("vx_profile")
        max_bins = max((int(v.bins) for v in self.vx_profiles), default=0)
        vx_headers = ["step", "bins", "x_window", *[f"vx_mean_{i}" for i in range(max_bins)]]
        ws_vx.append(vx_headers)
        for v in self.vx_profiles:
            vx = list(v.vx_mean)
            vx = vx[:max_bins] + [float("nan")] * max(0, max_bins - len(vx))
            ws_vx.append([int(v.step), int(v.bins), _safe_float(v.x_window), *[_safe_float(x) for x in vx]])
        ws_vx.freeze_panes = "A2"
        _autosize_worksheet(ws_vx)

        # ---- meta sheet
        ws_meta = wb.create_sheet("meta")
        ws_meta.append(["key", "value"])
        # Keep deterministic order for easier diffing.
        for k in sorted(self.meta.keys()):
            v = self.meta[k]
            if isinstance(v, (dict, list)):
                vv = json.dumps(v, sort_keys=True)
            elif v is None:
                vv = ""
            else:
                vv = str(v)
            ws_meta.append([str(k), vv])
        ws_meta.freeze_panes = "A2"
        _autosize_worksheet(ws_meta, max_width=80)

        wb.save(path)


def _autosize_worksheet(ws, *, max_width: int = 50) -> None:
    """
    Rough column autosizing (good enough for debugging exports).
    """
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                s = "nan"
            else:
                s = str(val)
            max_len = max(max_len, len(s))
        ws.column_dimensions[letter].width = min(max_width, max(10, max_len + 2))


